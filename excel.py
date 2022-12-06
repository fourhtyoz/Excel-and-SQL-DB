import sqlite3
from sqlite3 import Error
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Connects to DB
def create_connection(db_file):
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print(f'Connection established\nSQLite: {sqlite3.version}')
    except Error as e:
        print(e)
    return conn

# Main settings
db = 'main.db'
df = 'Данные.xlsx'
conn = create_connection(db)
c = conn.cursor()

# Find all cars
data = pd.read_excel(df)
cars_in_excel = data['Вагон №'].values
cars = set()
for car in cars_in_excel:
    cars.add(car)
cars = list(cars)

# Exports data from Excel to DB
def load_excel_data(conn, excel_file):
    try:
        data = pd.read_excel(excel_file)
        data.to_sql(name='main', con=conn)
        conn.commit()
    except FileNotFoundError as e:
        print(f'File not found. Error {e}')
    else:
        print(f'Data from {excel_file} loaded successfully')

# Helper function: Adds schemas to DB
def create_table(conn, table_name, schema_num):
    try:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS {name}(
            Рейс INTEGER PRIMARY KEY NOT NULL, 
            schema INTEGER DEFAULT {num}, 
            Вагон INTEGER NOT NULL, 
            Погрузка TEXT NOT NULL, 
            ГружПор TEXT NOT NULL)'''.format(name=table_name, num=schema_num))
        conn.commit()
    except Error as e:
        print(f'Create Table Error: {e}')
    else:
        print(f'Schema {schema_num} added to DB successfully')

# Main function: Adds schemas to DB
def car_tables(excel_file):
    data = pd.read_excel(excel_file)
    cars_in_excel = data['Вагон №'].values
    cars = set()
    for car in cars_in_excel:
        cars.add(car)
    count = 1
    for car in cars:
        create_table(conn, f'n_{car}', count)
        count += 1

# Adds data to car tables in DB
def add_subtables(conn, cars):
    try:
        c = conn.cursor()
        c.execute('SELECT * FROM main')
        result = c.fetchall()
        for row in result:
            for car in cars:
                if car in row: 
                    car_num, date_num, gruzh = (row[2], str(row[1][:10]), str(row[3]))
                    params = (car_num, date_num, gruzh)
                    # print(car_num, date_num, gruzh)
                    c.execute(f"INSERT INTO n_{car_num} (Рейс, Вагон, Погрузка, ГружПор) VALUES (NULL, ?, ?, ?)", params)
                    conn.commit()
    except Error as e:
        print(f'Add subtables: {e}')
    else:
        print('Data added to car tables successfully')

# Exports car tables from DB to Excel
def export_to_excel(conn, cars, excel_file):
    try:
        for car in cars:
            table = pd.read_sql(f"SELECT * FROM n_{car}", conn)
            frame = pd.DataFrame(table)
            wb = load_workbook(excel_file)
            wb.create_sheet(f'n_{car}')
            sheet = wb[f'n_{car}']
            for row in dataframe_to_rows(frame, index=False):
                sheet.append(row)
            wb.save(excel_file)
            wb.close()
    except Error as e:
        print(f'Exporting to Excel: {e}')
    else:
        print('Car tables exported to Excel successfully')

# Calculates revenue in each schema
def calculate_total(conn, cars, excel_file):
    try:
        c = conn.cursor()
        c.execute("SELECT * FROM main")
        result = c.fetchall()
        for car in cars:
            income = 0
            for row in result:
                if car in row: 
                    revenue, expenses = (row[6], row[11])
                    income += (revenue - expenses)
            wb = load_workbook(excel_file)
            ws = wb[f'n_{car}']
            ws['A12'].value = 'Общая доходность схемы:'
            ws['B12'].value = income
            wb.save(excel_file)
            wb.close()
    except Error as e:
        print(f'Calculating revenue: {e}')
    else:
        print('Total calculated successfully for each schema')

# Step 1
# Creates DB
load_excel_data(conn, df)

# Step 2
# Creates car schemas
car_tables(df)

# Step 3
# Creates subtables
add_subtables(conn, cars)

# Step 4
# Exports subtables to excel
export_to_excel(conn, cars, df)

# Step 5
# Calculates revenue for each sсhema
calculate_total(conn, cars, df)